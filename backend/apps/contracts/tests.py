# =============================================================================
# === backend/apps/contracts/tests.py ===
# =============================================================================
import io
from datetime import date, timedelta
from decimal import Decimal

import openpyxl
from apps.authentication.models import CustomUser
from apps.organizations.models import Organization, OrganizationMembership
from apps.service.models import Customer, Vehicle
from django.test import TestCase
from rest_framework.test import APITestCase

from .models import (Contract, ContractImport, ContractLineItem,
                     ContractVehicle, TerminPeriod)
from .parsing import (ContractParseError, diff_against_contract,
                      parse_hps_workbook, parse_rupiah)


def _build_test_workbook(rows):
    """Small helper — builds an in-memory .xlsx matching the real HPS
    template's structure, so tests read like the actual document
    rather than an abstract fixture."""
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


HEADER_ROWS = [
    ["PEKERJAAN : PENGADAAN PEMELIHARAAN DAN PERAWATAN..."],
    ["[REDACTED REFERENCE LINE]"],
    [],
    ["NO", "ITEM PEKERJAAN", "VOL", "SATUAN", "HARGA SATUAN", "JUMLAH"],
]


class ParseRupiahTests(TestCase):
    """
    Proves the actual real-world inconsistencies found in the one
    real HPS document reviewed for this project parse correctly —
    not just the clean, ideal case.
    """

    def test_plain_rp_format(self):
        self.assertEqual(parse_rupiah("Rp 200.000"), Decimal("200000"))

    def test_real_typo_rp_with_period(self):
        """The actual typo found in Group I, row 9 of the real Polda
        Kepri document: 'Rp.' instead of 'Rp'."""
        self.assertEqual(parse_rupiah("Rp. 4.000.000"), Decimal("4000000"))

    def test_numeric_cell_value(self):
        self.assertEqual(parse_rupiah(21600000), Decimal("21600000"))

    def test_comma_decimal_defensive_case(self):
        """Never observed in the real document, but Indonesian
        formatting could in principle use a comma decimal — must not
        be confused with '.' as a thousands separator."""
        self.assertEqual(parse_rupiah("Rp 1.234,50"), Decimal("1234.50"))

    def test_empty_value_raises(self):
        with self.assertRaises(ContractParseError):
            parse_rupiah("")

    def test_garbage_value_raises(self):
        with self.assertRaises(ContractParseError):
            parse_rupiah("tidak ada harga")


class ParseHpsWorkbookTests(TestCase):
    """
    Built against the VISIBLE structure of the real Polda Kepri HPS
    document (seen as PDF text, not yet the real .xlsx) — see
    parsing.py's own module docstring for why this still needs
    calibration against the real file.
    """

    def test_parses_two_vehicle_groups_with_the_real_typo(self):
        buf = _build_test_workbook(HEADER_ROWS + [
            ["I", "HYUNDAY TUCSON (9-XXXI)", 1, "MOBIL", "", "Rp 21.600.000"],
            [1, "Oli Mesin", 15, "Liter", "Rp 200.000", "Rp 3.000.000"],
            [2, "Filter Oli", 3, "Unit", "Rp 200.000", "Rp 600.000"],
            [9, "Perbaikan kaki-kaki", 1, "Paket", "Rp 4.000.000", "Rp. 4.000.000"],
            ["II", "TOYOTA AVANZA (921-XXXI)", 1, "MOBIL", "", "Rp 21.600.000"],
            [1, "Oli Mesin semi/ full synthetic", 12, "Liter", "Rp 200.000", "Rp 2.400.000"],
            [2, "Filter Oli", 3, "Unit", "Rp 120.000", "Rp 360.000"],
            ["", "TOTAL KESELURUHAN", "", "", "", "Rp 10.360.000"],
        ])
        parsed = parse_hps_workbook(buf)

        self.assertEqual(len(parsed.vehicle_groups), 2)
        g1, g2 = parsed.vehicle_groups
        self.assertEqual(g1.vehicle_model, "HYUNDAY TUCSON")
        self.assertEqual(g1.fleet_code, "9-XXXI")
        self.assertEqual(len(g1.line_items), 3)
        # The typo row parsed correctly despite "Rp." instead of "Rp":
        self.assertEqual(g1.line_items[2].subtotal, Decimal("4000000"))
        self.assertEqual(g2.fleet_code, "921-XXXI")
        self.assertEqual(parsed.document_total, Decimal("10360000"))
        self.assertEqual(parsed.computed_total, Decimal("10360000"))

    def test_missing_header_raises_parse_error(self):
        buf = _build_test_workbook([["not", "a", "real", "template"]])
        with self.assertRaises(ContractParseError):
            parse_hps_workbook(buf)

    def test_line_item_before_any_group_header_raises(self):
        """The template assumption genuinely doesn't hold for this
        file — must fail loudly, not silently attach the row to
        nothing."""
        buf = _build_test_workbook(HEADER_ROWS + [
            [1, "Oli Mesin", 15, "Liter", "Rp 200.000", "Rp 3.000.000"],
        ])
        with self.assertRaises(ContractParseError):
            parse_hps_workbook(buf)

    def test_no_vehicle_groups_found_raises(self):
        buf = _build_test_workbook(HEADER_ROWS)
        with self.assertRaises(ContractParseError):
            parse_hps_workbook(buf)

    def test_vehicle_name_without_fleet_code_falls_back_gracefully(self):
        """A malformed vehicle name shouldn't block importing every
        other vehicle in the same file — surfaced as an empty
        fleet_code for a human to fill in during review, not a hard
        failure."""
        buf = _build_test_workbook(HEADER_ROWS + [
            ["I", "SOME VEHICLE WITHOUT A CODE", 1, "MOBIL", "", "Rp 1.000.000"],
            [1, "Oli Mesin", 1, "Liter", "Rp 1.000.000", "Rp 1.000.000"],
        ])
        parsed = parse_hps_workbook(buf)
        self.assertEqual(parsed.vehicle_groups[0].fleet_code, "")
        self.assertEqual(parsed.vehicle_groups[0].vehicle_model, "SOME VEHICLE WITHOUT A CODE")


class ContractsAPITestBase(APITestCase):

    def setUp(self):
        self.org = Organization.objects.create(name="Arya Motor", invoice_code="AM")
        self.owner = CustomUser.objects.create_user(
            email="owner.contracts@test.id", password="pass12345!",
            full_name="Made Owner", role=CustomUser.Role.OWNER,
        )
        OrganizationMembership.objects.create(
            organization=self.org, user=self.owner, role="owner", is_active=True,
        )
        self.institutional_customer = Customer.objects.create(
            organization=self.org, name="Ditreskrimum & Dittahti Polda Kepri",
        )
        self.contract = Contract.objects.create(
            organization=self.org, customer=self.institutional_customer,
            title="Pengadaan Pemeliharaan Kendaraan R4/R6", fiscal_year=2026,
            termin_count=4,
        )
        self.client.force_authenticate(user=self.owner)


class DiffAgainstContractTests(ContractsAPITestBase):
    """
    Proves the diff engine correctly classifies every case using
    (fleet_code, source_row_no) as the matching key — not description
    text, which the real document already proves is unstable across
    different vehicles for conceptually-the-same job.
    """

    def setUp(self):
        super().setUp()
        self.vehicle = Vehicle.objects.create(
            organization=self.org, customer=self.institutional_customer,
            plate_number="9-XXXI", manufacture_year=2020,
            vehicle_type="Mobil", model="HYUNDAY TUCSON",
        )
        self.contract_vehicle = ContractVehicle.objects.create(
            organization=self.org, contract=self.contract,
            vehicle=self.vehicle, allocated_budget=Decimal("21600000"),
        )
        self.existing_item = ContractLineItem.objects.create(
            organization=self.org, contract_vehicle=self.contract_vehicle,
            source_row_no=1, description="Oli Mesin", volume=Decimal("15"),
            unit="Liter", unit_price=Decimal("200000"), subtotal=Decimal("3000000"),
        )

    def _parsed(self, rows):
        buf = _build_test_workbook(HEADER_ROWS + rows)
        return parse_hps_workbook(buf)

    def test_unchanged_line_is_not_flagged(self):
        parsed = self._parsed([
            ["I", "HYUNDAY TUCSON (9-XXXI)", 1, "MOBIL", "", "Rp 21.600.000"],
            [1, "Oli Mesin", 15, "Liter", "Rp 200.000", "Rp 3.000.000"],
            ["", "TOTAL KESELURUHAN", "", "", "", "Rp 3.000.000"],
        ])
        diff = diff_against_contract(parsed, self.contract)
        self.assertEqual(diff["unchanged_count"], 1)
        self.assertEqual(diff["added_items"], [])
        self.assertEqual(diff["changed_items"], [])
        self.assertEqual(diff["removed_items"], [])

    def test_changed_price_is_flagged_with_old_and_new(self):
        parsed = self._parsed([
            ["I", "HYUNDAY TUCSON (9-XXXI)", 1, "MOBIL", "", "Rp 21.600.000"],
            [1, "Oli Mesin", 15, "Liter", "Rp 250.000", "Rp 3.750.000"],  # price went up
            ["", "TOTAL KESELURUHAN", "", "", "", "Rp 3.750.000"],
        ])
        diff = diff_against_contract(parsed, self.contract)
        self.assertEqual(len(diff["changed_items"]), 1)
        change = diff["changed_items"][0]
        # existing_li.unit_price/subtotal are read back from real
        # DecimalField(decimal_places=2) columns — Django always
        # returns these quantized to two decimal places, so str()
        # faithfully produces "200000.00", not "200000". The diff
        # function's own str(existing_li.unit_price) call is correct;
        # this assertion was just wrong about what that produces.
        self.assertEqual(change["old"]["unit_price"], "200000.00")
        self.assertEqual(change["new"]["unit_price"], "250000")

    def test_new_line_on_existing_vehicle_is_added_item(self):
        parsed = self._parsed([
            ["I", "HYUNDAY TUCSON (9-XXXI)", 1, "MOBIL", "", "Rp 24.600.000"],
            [1, "Oli Mesin", 15, "Liter", "Rp 200.000", "Rp 3.000.000"],
            [2, "Filter Oli", 3, "Unit", "Rp 200.000", "Rp 600.000"],  # brand new row
            ["", "TOTAL KESELURUHAN", "", "", "", "Rp 3.600.000"],
        ])
        diff = diff_against_contract(parsed, self.contract)
        self.assertEqual(len(diff["added_items"]), 1)
        self.assertEqual(diff["added_items"][0]["row_no"], 2)

    def test_missing_line_is_removed_item(self):
        """Simulates a revision that drops row 1 entirely."""
        parsed = self._parsed([
            ["I", "HYUNDAY TUCSON (9-XXXI)", 1, "MOBIL", "", "Rp 0"],
            ["", "TOTAL KESELURUHAN", "", "", "", "Rp 0"],
        ])
        diff = diff_against_contract(parsed, self.contract)
        self.assertEqual(len(diff["removed_items"]), 1)
        self.assertEqual(diff["removed_items"][0]["row_no"], 1)

    def test_brand_new_vehicle_is_added_vehicle_not_added_item(self):
        parsed = self._parsed([
            ["I", "HYUNDAY TUCSON (9-XXXI)", 1, "MOBIL", "", "Rp 21.600.000"],
            [1, "Oli Mesin", 15, "Liter", "Rp 200.000", "Rp 3.000.000"],
            ["II", "TOYOTA AVANZA (921-XXXI)", 1, "MOBIL", "", "Rp 2.400.000"],
            [1, "Oli Mesin", 12, "Liter", "Rp 200.000", "Rp 2.400.000"],
            ["", "TOTAL KESELURUHAN", "", "", "", "Rp 5.400.000"],
        ])
        diff = diff_against_contract(parsed, self.contract)
        self.assertEqual(len(diff["added_vehicles"]), 1)
        self.assertEqual(diff["added_vehicles"][0]["fleet_code"], "921-XXXI")
        # Existing vehicle's own unchanged line must not also appear
        # as noise elsewhere:
        self.assertEqual(diff["unchanged_count"], 1)


class ContractImportApplyTests(ContractsAPITestBase):
    """
    Proves the actual mutation logic — the part with real
    consequences if it gets the ordering or the supersede logic
    wrong.
    """

    def test_apply_creates_vehicle_and_line_items_for_new_contract(self):
        """The very first import for a brand-new Contract — no
        separate 'create from scratch' path, just a diff where
        everything happens to be added."""
        contract_import = ContractImport.objects.create(
            organization=self.org, contract=self.contract,
            original_file="contract_imports/test.xlsx", uploaded_by=self.owner,
        )
        confirmed_diff = {
            "added_vehicles": [{
                "fleet_code": "9-XXXI",
                "vehicle_model": "HYUNDAY TUCSON",
                "manufacture_year": 2020,  # filled in by the reviewer —
                # never present in the source document itself
                "vehicle_type": "Mobil",
                "allocated_budget": "21600000",
                "line_items": [
                    {"row_no": 1, "description": "Oli Mesin", "volume": "15",
                     "unit": "Liter", "unit_price": "200000", "subtotal": "3000000"},
                ],
            }],
        }

        contract_import.apply(confirmed_diff, applied_by=self.owner)

        self.assertEqual(Vehicle.objects.filter(plate_number="9-XXXI").count(), 1)
        vehicle = Vehicle.objects.get(plate_number="9-XXXI")
        self.assertEqual(vehicle.manufacture_year, 2020)
        cv = ContractVehicle.objects.get(contract=self.contract, vehicle=vehicle)
        self.assertEqual(cv.allocated_budget, Decimal("21600000"))
        line_item = ContractLineItem.objects.get(contract_vehicle=cv, source_row_no=1)
        self.assertEqual(line_item.status, "ACTIVE")
        self.assertEqual(line_item.subtotal, Decimal("3000000"))

        contract_import.refresh_from_db()
        self.assertEqual(contract_import.status, "APPLIED")
        self.assertEqual(contract_import.applied_by, self.owner)

    def test_apply_reuses_existing_vehicle_instead_of_creating_duplicate(self):
        """
        The real bug this test exists to prevent regressing: the same
        real fleet vehicle reappearing in a second contract (the
        exact scenario ContractVehicle was built as its own join
        table for) used to crash outright — Vehicle.objects.create()
        colliding with the (organization, plate_number) unique
        constraint, since apply() never checked whether a Vehicle
        with that plate already existed anywhere in the org before
        trying to create one.
        """
        existing_vehicle = Vehicle.objects.create(
            organization=self.org, customer=self.institutional_customer,
            plate_number="2-XXXI", manufacture_year=2018,
            vehicle_type="Mobil", model="HONDA BRIO",
        )
        # A second, separate Contract — the same fleet vehicle
        # showing up in a different fiscal year's tender, same
        # organization.
        second_contract = Contract.objects.create(
            organization=self.org, customer=self.institutional_customer,
            title="Pengadaan Tahun Kedua", fiscal_year=2027, termin_count=4,
        )
        contract_import = ContractImport.objects.create(
            organization=self.org, contract=second_contract,
            original_file="contract_imports/year2.xlsx", uploaded_by=self.owner,
        )
        confirmed_diff = {
            "added_vehicles": [{
                "fleet_code": "2-XXXI",
                "vehicle_model": "HONDA BRIO",
                "existing_vehicle_id": str(existing_vehicle.id),
                # Deliberately no manufacture_year/vehicle_type here —
                # a reuse-case entry never needs them, per the review
                # screen's own logic.
                "allocated_budget": "19700000",
                "line_items": [
                    {"row_no": 1, "description": "Oli Mesin", "volume": "15",
                     "unit": "Liter", "unit_price": "200000", "subtotal": "3000000"},
                ],
            }],
        }

        contract_import.apply(confirmed_diff, applied_by=self.owner)

        # No duplicate Vehicle created — still exactly one with this plate:
        self.assertEqual(Vehicle.objects.filter(plate_number="2-XXXI").count(), 1)
        vehicle = Vehicle.objects.get(plate_number="2-XXXI")
        self.assertEqual(vehicle.id, existing_vehicle.id)
        # Its original manufacture_year is untouched, not overwritten:
        self.assertEqual(vehicle.manufacture_year, 2018)
        # Correctly linked to the NEW contract:
        cv = ContractVehicle.objects.get(contract=second_contract, vehicle=vehicle)
        self.assertEqual(cv.allocated_budget, Decimal("19700000"))

    def test_diff_detects_reusable_vehicle_from_a_different_contract(self):
        """
        Proves diff_against_contract() itself surfaces
        existing_vehicle_id correctly — this is what makes the
        review screen able to skip asking for manufacture_year at
        all for a vehicle that already has one.
        """
        Vehicle.objects.create(
            organization=self.org, customer=self.institutional_customer,
            plate_number="2-XXXI", manufacture_year=2018,
            vehicle_type="Mobil", model="HONDA BRIO",
        )
        second_contract = Contract.objects.create(
            organization=self.org, customer=self.institutional_customer,
            title="Pengadaan Tahun Kedua", fiscal_year=2027, termin_count=4,
        )
        buf = _build_test_workbook(HEADER_ROWS + [
            ["", "HONDA BRIO (2-XXXI)", 1, "MOBIL", "", "Rp 19.700.000"],
            [1, "Oli Mesin", 15, "Liter", "Rp 200.000", "Rp 3.000.000"],
        ])
        parsed = parse_hps_workbook(buf)
        diff = diff_against_contract(parsed, second_contract)

        self.assertEqual(len(diff["added_vehicles"]), 1)
        entry = diff["added_vehicles"][0]
        self.assertIsNotNone(entry["existing_vehicle_id"])
        self.assertEqual(entry["existing_vehicle_model"], "HONDA BRIO")

    def test_apply_supersedes_old_line_never_hard_deletes(self):
        """The core safety property this whole design exists for: a
        real WorkOrder created earlier may already reference the old
        line item, so it must never vanish, only be marked
        SUPERSEDED with a pointer to what replaced it."""
        vehicle = Vehicle.objects.create(
            organization=self.org, customer=self.institutional_customer,
            plate_number="9-XXXI", manufacture_year=2020,
            vehicle_type="Mobil", model="HYUNDAY TUCSON",
        )
        cv = ContractVehicle.objects.create(
            organization=self.org, contract=self.contract,
            vehicle=vehicle, allocated_budget=Decimal("21600000"),
        )
        old_item = ContractLineItem.objects.create(
            organization=self.org, contract_vehicle=cv, source_row_no=1,
            description="Oli Mesin", volume=Decimal("15"), unit="Liter",
            unit_price=Decimal("200000"), subtotal=Decimal("3000000"),
        )

        contract_import = ContractImport.objects.create(
            organization=self.org, contract=self.contract,
            original_file="contract_imports/revision.xlsx", uploaded_by=self.owner,
        )
        confirmed_diff = {
            "changed_items": [{
                "fleet_code": "9-XXXI", "row_no": 1,
                "old": {"description": "Oli Mesin", "volume": "15", "unit": "Liter",
                        "unit_price": "200000", "subtotal": "3000000"},
                "new": {"description": "Oli Mesin", "volume": "15", "unit": "Liter",
                        "unit_price": "250000", "subtotal": "3750000"},
            }],
        }
        contract_import.apply(confirmed_diff, applied_by=self.owner)

        old_item.refresh_from_db()
        self.assertEqual(old_item.status, "SUPERSEDED")
        self.assertIsNotNone(old_item.superseded_by)

        new_item = old_item.superseded_by
        self.assertEqual(new_item.status, "ACTIVE")
        self.assertEqual(new_item.unit_price, Decimal("250000"))

        # Only ever one ACTIVE row at this exact position — the real
        # guarantee the UniqueConstraint on the model enforces at the
        # DB level, confirmed here at the application level too.
        active_count = ContractLineItem.objects.filter(
            contract_vehicle=cv, source_row_no=1, status="ACTIVE",
        ).count()
        self.assertEqual(active_count, 1)

    def test_cannot_apply_an_already_applied_import(self):
        contract_import = ContractImport.objects.create(
            organization=self.org, contract=self.contract,
            original_file="contract_imports/test.xlsx", uploaded_by=self.owner,
            status="APPLIED",
        )
        with self.assertRaises(ValueError):
            contract_import.apply({}, applied_by=self.owner)

    def test_reject_marks_rejected_without_touching_line_items(self):
        contract_import = ContractImport.objects.create(
            organization=self.org, contract=self.contract,
            original_file="contract_imports/test.xlsx", uploaded_by=self.owner,
        )
        contract_import.reject()
        contract_import.refresh_from_db()
        self.assertEqual(contract_import.status, "REJECTED")
        self.assertEqual(ContractLineItem.objects.count(), 0)


class ContractsTenantIsolationTests(ContractsAPITestBase):

    def setUp(self):
        super().setUp()
        self.other_org = Organization.objects.create(name="Bengkel Lain Kontrak")
        self.other_owner = CustomUser.objects.create_user(
            email="owner.othercontracts@test.id", password="pass12345!",
            full_name="Other Owner", role=CustomUser.Role.OWNER,
        )
        OrganizationMembership.objects.create(
            organization=self.other_org, user=self.other_owner, role="owner", is_active=True,
        )

    def test_org_b_cannot_see_org_a_contracts(self):
        self.client.force_authenticate(user=self.other_owner)
        resp = self.client.get("/api/contracts/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["count"], 0)

    def test_org_b_cannot_retrieve_org_a_contract_detail(self):
        self.client.force_authenticate(user=self.other_owner)
        resp = self.client.get(f"/api/contracts/{self.contract.id}/")
        self.assertEqual(resp.status_code, 404)


class ContractImportListTests(ContractsAPITestBase):
    """
    Covers the GET side of the nested imports endpoint — the data
    source for a contract-detail page's "Riwayat Import" section.
    """

    def test_lists_imports_newest_first(self):
        older = ContractImport.objects.create(
            organization=self.org, contract=self.contract,
            original_file="contract_imports/first.xlsx", uploaded_by=self.owner,
            status="APPLIED",
        )
        newer = ContractImport.objects.create(
            organization=self.org, contract=self.contract,
            original_file="contract_imports/second.xlsx", uploaded_by=self.owner,
        )
        resp = self.client.get(f"/api/contracts/{self.contract.id}/imports/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["count"], 2)
        # Newest first, per ContractImport's own Meta.ordering:
        self.assertEqual(resp.data["results"][0]["id"], str(newer.id))
        self.assertEqual(resp.data["results"][1]["id"], str(older.id))

    def test_empty_list_for_contract_with_no_imports(self):
        resp = self.client.get(f"/api/contracts/{self.contract.id}/imports/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["count"], 0)

    def test_returns_404_for_nonexistent_contract(self):
        import uuid as uuid_module
        resp = self.client.get(f"/api/contracts/{uuid_module.uuid4()}/imports/")
        self.assertEqual(resp.status_code, 404)


class TerminPeriodGenerationTests(ContractsAPITestBase):
    """
    Made's own real termin-tracking example from the 28 Jul meeting
    (the Avanza 849 XXXI-28). Uses the real creation API, not direct
    ORM Contract.objects.create() — generate_termin_periods() is
    triggered by ContractListView.post(), by design, not by
    Contract.save() itself, so a test needs to go through the real
    endpoint to prove this actually happens on real contract creation.
    """

    def test_creating_a_4x_contract_generates_four_periods(self):
        resp = self.client.post("/api/contracts/", {
            "customer": str(self.institutional_customer.id),
            "title": "Pengadaan Tahun Baru", "fiscal_year": 2026,
            "termin_count": 4, "start_date": "2026-01-15",
        }, format="json")
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(len(resp.data["contract"]["termin_periods"]), 4)

    def test_creating_a_3x_contract_generates_three_periods(self):
        resp = self.client.post("/api/contracts/", {
            "customer": str(self.institutional_customer.id),
            "title": "Pengadaan Tahun Baru", "fiscal_year": 2026,
            "termin_count": 3, "start_date": "2026-01-15",
        }, format="json")
        self.assertEqual(resp.status_code, 201)
        self.assertEqual(len(resp.data["contract"]["termin_periods"]), 3)

    def test_4x_due_dates_are_every_3_months_from_start_date(self):
        """Made's own confirmed numbers: 4x/year is every 3 months —
        not approximate, an exact integer division of 12."""
        resp = self.client.post("/api/contracts/", {
            "customer": str(self.institutional_customer.id),
            "title": "Pengadaan Tahun Baru", "fiscal_year": 2026,
            "termin_count": 4, "start_date": "2026-01-15",
        }, format="json")
        due_dates = [p["jatuh_tempo"] for p in resp.data["contract"]["termin_periods"]]
        self.assertEqual(due_dates, ["2026-04-15", "2026-07-15", "2026-10-15", "2027-01-15"])

    def test_3x_due_dates_are_every_4_months_from_start_date(self):
        resp = self.client.post("/api/contracts/", {
            "customer": str(self.institutional_customer.id),
            "title": "Pengadaan Tahun Baru", "fiscal_year": 2026,
            "termin_count": 3, "start_date": "2026-01-15",
        }, format="json")
        due_dates = [p["jatuh_tempo"] for p in resp.data["contract"]["termin_periods"]]
        self.assertEqual(due_dates, ["2026-05-15", "2026-09-15", "2027-01-15"])

    def test_month_arithmetic_clamps_end_of_month_correctly(self):
        """
        31 Jan + 3 months should land on 30 Apr (not an invalid
        '31 Apr'), and 31 Jan + 6 months should land on 31 Jul (a
        real 31-day month) — proving the clamp only ever reduces the
        day when the target month genuinely has fewer days, not
        unconditionally.
        """
        resp = self.client.post("/api/contracts/", {
            "customer": str(self.institutional_customer.id),
            "title": "Pengadaan Akhir Bulan", "fiscal_year": 2026,
            "termin_count": 4, "start_date": "2026-01-31",
        }, format="json")
        due_dates = [p["jatuh_tempo"] for p in resp.data["contract"]["termin_periods"]]
        self.assertEqual(due_dates[0], "2026-04-30")
        self.assertEqual(due_dates[1], "2026-07-31")

    def test_amount_expected_starts_at_zero_before_any_import(self):
        """
        A brand-new Contract has zero ContractVehicles — there's no
        real budget to split yet. See Contract.generate_termin_
        periods()'s own docstring for why this is 0, not left null.
        """
        resp = self.client.post("/api/contracts/", {
            "customer": str(self.institutional_customer.id),
            "title": "Pengadaan Tahun Baru", "fiscal_year": 2026,
            "termin_count": 4, "start_date": "2026-01-15",
        }, format="json")
        amounts = [p["amount_expected"] for p in resp.data["contract"]["termin_periods"]]
        self.assertTrue(all(Decimal(a) == Decimal("0") for a in amounts))


class ContractGenerateTerminBackfillTests(ContractsAPITestBase):
    """
    The real gap this closes: self.contract (from the base fixture)
    is created via direct ORM Contract.objects.create(), never
    through ContractListView.post() — the exact same situation as
    every real Contract that existed before this feature shipped.
    Proven directly against Chris's own real screenshot: a genuine
    pre-existing contract showing zero termin periods, with no way
    to fix it until this endpoint existed.
    """

    def test_generates_periods_for_a_contract_that_predates_the_feature(self):
        self.assertEqual(self.contract.termin_periods.count(), 0)
        resp = self.client.post(f"/api/contracts/{self.contract.id}/generate-termin/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(len(resp.data["contract"]["termin_periods"]), 4)

    def test_rejects_with_409_if_periods_already_exist(self):
        """
        generate_termin_periods() itself isn't safe to call twice —
        this is the guard that actually prevents hitting that."""
        self.contract.generate_termin_periods()
        resp = self.client.post(f"/api/contracts/{self.contract.id}/generate-termin/")
        self.assertEqual(resp.status_code, 409)
        self.assertEqual(self.contract.termin_periods.count(), 4)  # unchanged, not duplicated


class TerminPeriodRecalculationTests(ContractsAPITestBase):
    """
    Proves the actual live-recalculation hook in ContractImport.
    apply() — the real reason amount_expected is NOT frozen the same
    way Invoice/InvoiceLineItem's own financial values are.
    """

    def setUp(self):
        super().setUp()
        self.contract.generate_termin_periods()

    def _apply_a_vehicle(self, fleet_code, budget):
        contract_import = ContractImport.objects.create(
            organization=self.org, contract=self.contract,
            original_file=f"contract_imports/{fleet_code}.xlsx", uploaded_by=self.owner,
        )
        confirmed_diff = {
            "added_vehicles": [{
                "fleet_code": fleet_code, "vehicle_model": "HYUNDAY TUCSON",
                "manufacture_year": 2020, "vehicle_type": "Mobil",
                "allocated_budget": budget,
                "line_items": [
                    {"row_no": 1, "description": "Oli Mesin", "volume": "15",
                     "unit": "Liter", "unit_price": "200000", "subtotal": "3000000"},
                ],
            }],
        }
        contract_import.apply(confirmed_diff, applied_by=self.owner)

    def test_amounts_split_evenly_across_all_unrealized_periods_after_apply(self):
        self._apply_a_vehicle("9-XXXI", "20000000")
        periods = list(self.contract.termin_periods.order_by("sequence"))
        self.assertEqual(len(periods), 4)
        for p in periods:
            p.refresh_from_db()
            self.assertEqual(p.amount_expected, Decimal("5000000.00"))

    def test_recalculation_runs_again_on_a_second_apply_reflecting_new_total(self):
        """The real reason this stays live-recalculable rather than a
        one-time snapshot: a contract amendment adding more vehicles
        via a second import should update the expectation, not leave
        it stuck at the original figure."""
        self._apply_a_vehicle("9-XXXI", "20000000")
        self._apply_a_vehicle("921-XXXI", "20000000")
        periods = list(self.contract.termin_periods.order_by("sequence"))
        for p in periods:
            p.refresh_from_db()
            self.assertEqual(p.amount_expected, Decimal("10000000.00"))

    def test_already_realized_period_is_excluded_from_recalculation(self):
        """
        The one thing that IS frozen forever: a termin already marked
        received must never have its own expected figure rewritten
        after the fact, regardless of what the contract's total
        budget does later.
        """
        first_period = self.contract.termin_periods.order_by("sequence").first()
        first_period.record_realization(Decimal("999999"), received_date=date(2026, 3, 1))

        self._apply_a_vehicle("9-XXXI", "20000000")

        first_period.refresh_from_db()
        self.assertEqual(first_period.amount_expected, Decimal("0"))  # untouched
        self.assertEqual(first_period.amount_received, Decimal("999999"))

        other_periods = self.contract.termin_periods.exclude(id=first_period.id)
        for p in other_periods:
            p.refresh_from_db()
            # Split across the 3 remaining unrealized periods, not 4:
            self.assertEqual(p.amount_expected, Decimal("6666666.67"))


class TerminPeriodRealizeAPITests(ContractsAPITestBase):

    def setUp(self):
        super().setUp()
        self.contract.generate_termin_periods()
        self.period = self.contract.termin_periods.order_by("sequence").first()

    def test_realize_sets_amount_and_defaults_received_at_to_today(self):
        resp = self.client.post(
            f"/api/termin-periods/{self.period.id}/realize/",
            {"amount_received": "5000000"}, format="json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertTrue(resp.data["termin_period"]["is_realized"])
        self.assertEqual(resp.data["termin_period"]["received_at"], date.today().isoformat())

    def test_realize_accepts_an_explicit_received_date(self):
        resp = self.client.post(
            f"/api/termin-periods/{self.period.id}/realize/",
            {"amount_received": "5000000", "received_at": "2026-03-15"}, format="json",
        )
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.data["termin_period"]["received_at"], "2026-03-15")

    def test_realize_rejects_missing_amount(self):
        resp = self.client.post(f"/api/termin-periods/{self.period.id}/realize/", {}, format="json")
        self.assertEqual(resp.status_code, 400)

    def test_realize_rejects_a_genuinely_invalid_amount_with_400_not_500(self):
        """
        The exact bug caught and fixed before shipping: Decimal("not-
        a-number") raises decimal.InvalidOperation, which is neither
        a ValueError nor a TypeError — an earlier version of this
        view's except clause would have let this surface as an
        unhandled 500 instead of a clean 400.
        """
        resp = self.client.post(
            f"/api/termin-periods/{self.period.id}/realize/",
            {"amount_received": "not-a-number"}, format="json",
        )
        self.assertEqual(resp.status_code, 400)


class TerminPeriodModelTests(ContractsAPITestBase):

    def setUp(self):
        super().setUp()
        self.contract.generate_termin_periods()

    def test_is_realized_false_until_recorded(self):
        period = self.contract.termin_periods.first()
        self.assertFalse(period.is_realized)

    def test_is_overdue_true_when_due_date_passed_and_not_realized(self):
        period = self.contract.termin_periods.first()
        period.jatuh_tempo = date.today() - timedelta(days=1)
        period.save(update_fields=["jatuh_tempo"])
        self.assertTrue(period.is_overdue)

    def test_is_overdue_false_once_realized_even_if_late(self):
        period = self.contract.termin_periods.first()
        period.jatuh_tempo = date.today() - timedelta(days=10)
        period.save(update_fields=["jatuh_tempo"])
        period.record_realization(Decimal("5000000"))
        self.assertFalse(period.is_overdue)

    def test_is_overdue_false_before_due_date(self):
        period = self.contract.termin_periods.first()
        period.jatuh_tempo = date.today() + timedelta(days=30)
        period.save(update_fields=["jatuh_tempo"])
        self.assertFalse(period.is_overdue)


class ContractExportTerminTests(ContractsAPITestBase):
    """
    Made's own ask, 28 Jul meeting: export to Word/Excel for sending
    to institutions. Proves real cell values in the actual generated
    file, not just that the endpoint returns 200 — same discipline
    already proven on parsing.py's own read-side tests, applied here
    to the write side.
    """

    def setUp(self):
        super().setUp()
        self.contract.generate_termin_periods()

    def _get_workbook(self, contract_id):
        resp = self.client.get(f"/api/contracts/{contract_id}/export-termin/")
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(
            resp["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        # A plain HttpResponse, not StreamingHttpResponse — resp.content
        # is always the full body here, no need to branch on it.
        return openpyxl.load_workbook(io.BytesIO(resp.content))

    def test_export_contains_real_contract_header_info(self):
        wb = self._get_workbook(self.contract.id)
        ws = wb.active
        self.assertEqual(ws["B3"].value, self.institutional_customer.name)
        self.assertEqual(ws["B4"].value, self.contract.title)
        self.assertEqual(ws["B5"].value, self.contract.fiscal_year)

    def test_export_contains_real_termin_row_values(self):
        period = self.contract.termin_periods.first()
        period.record_realization(Decimal("5000000.00"), received_date=date(2026, 3, 1))

        wb = self._get_workbook(self.contract.id)
        ws = wb.active
        # Row 9 is the first data row — table header sits at row 8,
        # per build_termin_report_workbook's own layout.
        self.assertEqual(ws.cell(row=9, column=1).value, 1)
        self.assertEqual(Decimal(str(ws.cell(row=9, column=3).value)), Decimal("0"))  # amount_expected, pre-import
        self.assertEqual(Decimal(str(ws.cell(row=9, column=4).value)), Decimal("5000000"))
        self.assertEqual(ws.cell(row=9, column=6).value, "Direalisasi")

    def test_export_totals_row_sums_correctly(self):
        """
        Proves the totals aren't just re-displaying one period's
        figures — genuinely summed across all of them, verified
        against real Decimal arithmetic, not assumed correct.
        """
        periods = list(self.contract.termin_periods.order_by("sequence"))
        periods[0].record_realization(Decimal("3000000"))
        periods[1].record_realization(Decimal("2000000"))

        wb = self._get_workbook(self.contract.id)
        ws = wb.active
        # 4 periods + header row (8) + 1 = totals lands on row 13.
        total_received = Decimal(str(ws.cell(row=13, column=4).value))
        self.assertEqual(total_received, Decimal("5000000"))

    def test_export_filename_sanitizes_unsafe_characters(self):
        """
        The exact real-world case this project's own test data
        already surfaces: a contract title containing a slash
        ("...Kendaraan R4/R6") would otherwise land in a filename a
        real filesystem could choke on.
        """
        contract = Contract.objects.create(
            organization=self.org, customer=self.institutional_customer,
            title="Pengadaan Pemeliharaan Kendaraan R4/R6", fiscal_year=2026, termin_count=4,
        )
        contract.generate_termin_periods()
        resp = self.client.get(f"/api/contracts/{contract.id}/export-termin/")
        disposition = resp["Content-Disposition"]
        self.assertNotIn("/", disposition.split("filename=")[1])
