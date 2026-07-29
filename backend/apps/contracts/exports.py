# =============================================================================
# === backend/apps/contracts/exports.py ===
# =============================================================================
"""
Arthasee — Contracts, Termin Report Export

Own module, not stuffed into models.py — mirrors parsing.py's own
precedent directly: import logic already lives separately from the
models it produces, so export logic (the reverse direction) gets the
same treatment. Reuses openpyxl deliberately — already a proven,
tested dependency in this project (the same library parsing.py uses
to READ real HPS files), not a new one (python-docx) for a Word
document Made hasn't specifically asked for over Excel.

This needs to actually look like a real document someone would send
to Polresta/Polsek, not a raw data dump — real header info, styled
table, a totals row, real Rupiah number formatting on the numeric
cells (not pre-formatted strings, so the recipient can still use
these as real numbers in their own spreadsheet).
"""
from decimal import Decimal

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

STATUS_LABEL = {
    "REALIZED": "Direalisasi",
    "OVERDUE":  "Jatuh Tempo",
    "PENDING":  "Belum Jatuh Tempo",
}

INDONESIAN_MONTHS = [
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
]


def _format_date_id(d):
    """
    Plain, dependency-free Indonesian date formatting — deliberately
    NOT strftime("%B"), which depends on Python's OS-level locale
    actually having "id_ID" installed on whatever server this runs
    on. That's not guaranteed, and on a system without it, strftime
    silently falls back to English month names rather than raising
    an error — exactly the bug caught in a real exported file (see
    the conversation this was fixed from: real dates coming back as
    "29 October 2026" instead of "29 Oktober 2026"). Also
    deliberately not Django's own format-localization system — untested
    here, and this project has been careful throughout not to trust
    a framework layer's behavior without verifying it directly. A
    static lookup table costs nothing and can never silently produce
    the wrong language.
    """
    return f"{d.day} {INDONESIAN_MONTHS[d.month - 1]} {d.year}"


def _termin_status_label(period):
    if period.is_realized:
        return STATUS_LABEL["REALIZED"]
    if period.is_overdue:
        return STATUS_LABEL["OVERDUE"]
    return STATUS_LABEL["PENDING"]


def build_termin_report_workbook(contract):
    """
    Returns a real openpyxl.Workbook — the caller (the view) is
    responsible for writing it to a buffer and returning it as an
    actual file download. Kept as a pure function returning a
    Workbook object, not writing to disk/response itself, so this
    stays independently testable — a test can build the workbook,
    save it to an in-memory buffer, and re-open it with openpyxl to
    verify real cell values, the same discipline already proven on
    parsing.py's own read-side tests.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Termin"

    bold_font   = Font(bold=True)
    title_font  = Font(bold=True, size=14)
    header_fill = PatternFill(start_color="2E2E2E", end_color="2E2E2E", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    thin_side   = Side(style="thin", color="CCCCCC")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    rupiah_format = "#,##0"

    # ── Header block ──────────────────────────────────────────────
    ws["A1"] = "LAPORAN STATUS TERMIN PEMBAYARAN"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")

    header_fields = [
        ("Klien", contract.customer.name),
        ("Judul Pekerjaan", contract.title),
        ("Tahun Anggaran", contract.fiscal_year),
        ("Jumlah Termin", f"{contract.termin_count}x per tahun"),
    ]
    for offset, (label, value) in enumerate(header_fields):
        row = 3 + offset
        ws.cell(row=row, column=1, value=label).font = bold_font
        ws.cell(row=row, column=2, value=value)

    # ── Table header ──────────────────────────────────────────────
    table_header_row = 8
    columns = ["Termin Ke-", "Jatuh Tempo", "Perkiraan (Rp)", "Realisasi (Rp)", "Tanggal Diterima", "Status"]
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=table_header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # ── Table rows ────────────────────────────────────────────────
    periods = contract.termin_periods.order_by("sequence")
    total_expected = Decimal("0")
    total_received = Decimal("0")

    row = table_header_row + 1
    for period in periods:
        ws.cell(row=row, column=1, value=period.sequence)
        ws.cell(row=row, column=2, value=_format_date_id(period.jatuh_tempo))

        expected_cell = ws.cell(row=row, column=3, value=period.amount_expected)
        expected_cell.number_format = rupiah_format

        received_cell = ws.cell(row=row, column=4, value=period.amount_received)
        received_cell.number_format = rupiah_format

        ws.cell(row=row, column=5, value=_format_date_id(period.received_at) if period.received_at else "—")
        ws.cell(row=row, column=6, value=_termin_status_label(period))

        for col_idx in range(1, 7):
            ws.cell(row=row, column=col_idx).border = thin_border

        total_expected += period.amount_expected
        if period.amount_received is not None:
            total_received += period.amount_received
        row += 1

    # ── Totals row ────────────────────────────────────────────────
    ws.cell(row=row, column=1, value="Total").font = bold_font
    total_expected_cell = ws.cell(row=row, column=3, value=total_expected)
    total_expected_cell.font = bold_font
    total_expected_cell.number_format = rupiah_format
    total_received_cell = ws.cell(row=row, column=4, value=total_received)
    total_received_cell.font = bold_font
    total_received_cell.number_format = rupiah_format

    # ── Column widths ─────────────────────────────────────────────
    for idx, width in enumerate([12, 18, 18, 18, 18, 20], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    return wb
