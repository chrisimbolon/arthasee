# =============================================================================
# === backend/apps/contracts/parsing.py ===
# =============================================================================
"""
The actual "engine" Made asked for: turns an HPS/RAB Excel workbook
into structured data apps.contracts can diff against a Contract's
existing live line items, and — once a human reviews that diff —
promote into real records.

IMPORTANT CAVEAT, read before touching this file: written against the
real HPS document reviewed for this project (Pengadaan Ditreskrimum
& Dittahti Polda Kepri, T.A. 2026) — but that document was only ever
seen as PDF text during design, never as the actual .xlsx file. The
column-detection and Roman-numeral group-detection logic below is
built from the VISIBLE STRUCTURE of that PDF, not verified against
real cell merges, styles, or formulas in an actual workbook. This
needs calibration against the real file before it touches production
data — see tests.py for exactly what was assumed, so a mismatch
against the real file shows up as a failing test, not a silent
misread in production.

Confirmed with Chris: Made uses one consistent template every time —
this is what makes a deterministic parser (as opposed to a fuzzy/
heuristic one) a reasonable engineering choice at all. If that ever
stops being true, this file is the first thing that needs to change.
"""
import re
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from typing import List, Optional

import openpyxl

FLEET_CODE_RE    = re.compile(r"\(([^)]+)\)\s*$")
RUPIAH_KEEP_RE   = re.compile(r"[^\d,]")


class ContractParseError(Exception):
    """
    Raised when the workbook doesn't match the expected template
    closely enough to parse safely. Always surfaced to the reviewer
    as ContractImport.parse_error — never silently guessed past. A
    wrong guess here means real budget numbers entering the system
    wrong, so failing loudly is the only acceptable behavior.
    """


@dataclass
class ParsedLineItem:
    row_no: int
    description: str
    volume: Decimal
    unit: str
    unit_price: Decimal
    subtotal: Decimal


@dataclass
class ParsedVehicleGroup:
    group_label: str            # "I", "II", ...
    raw_name: str                # "HYUNDAY TUCSON (9-XXXI)" as printed
    vehicle_model: str           # "HYUNDAY TUCSON"
    fleet_code: str              # "9-XXXI" — used as Vehicle.plate_number
    allocated_budget: Decimal
    line_items: List[ParsedLineItem] = field(default_factory=list)

    @property
    def computed_subtotal(self) -> Decimal:
        return sum((li.subtotal for li in self.line_items), Decimal("0"))


@dataclass
class ParsedContract:
    vehicle_groups: List[ParsedVehicleGroup]
    document_total: Optional[Decimal]

    @property
    def computed_total(self) -> Decimal:
        return sum((g.computed_subtotal for g in self.vehicle_groups), Decimal("0"))


def parse_rupiah(raw) -> Decimal:
    """
    Handles every variant actually seen in the one real document
    reviewed for this project: "Rp 200.000", and "Rp. 4.000.000" —
    the real typo found in Group I, row 9 of the Polda Kepri HPS,
    which is exactly the kind of small human inconsistency a naive
    parser breaks on. Also handles plain numeric cells (if the
    original workbook has HARGA SATUAN typed as a real number rather
    than text), and defensively — though never observed in the real
    document — a comma-decimal form.

    Indonesian number formatting uses "." as a THOUSANDS separator,
    not a decimal point (the opposite of US formatting) — so a "."
    here can never be safely treated as a decimal.
    """
    if raw is None or raw == "":
        raise ContractParseError("Empty value where a Rupiah amount was expected.")
    if isinstance(raw, (int, float, Decimal)):
        return Decimal(str(raw))

    text = str(raw).strip()
    # Strip "Rp" / "Rp." — not anchored to the string start, since the
    # real typo ("Rp." with the period directly after) must still
    # match this.
    text = re.sub(r"Rp\.?", "", text, flags=re.IGNORECASE).strip()
    text = RUPIAH_KEEP_RE.sub("", text)  # now only digits and commas remain

    if "," in text:
        integer_part, _, decimal_part = text.rpartition(",")
        integer_part = integer_part.replace(".", "")
        try:
            return Decimal(f"{integer_part}.{decimal_part}")
        except InvalidOperation:
            raise ContractParseError(f"Could not parse Rupiah amount: {raw!r}")

    text = text.replace(".", "")
    if not text.isdigit():
        raise ContractParseError(f"Could not parse Rupiah amount: {raw!r}")
    return Decimal(text)


def _is_group_header_row(satuan_cell) -> bool:
    """
    The real, structural signal for "this row starts a new vehicle."
    Originally required a Roman numeral (I, II, III...) in the NO
    column too, based on the one document seen as PDF text during
    design — but the real .xlsx (once actually opened) turned out to
    leave NO completely blank on its own group-header row, with only
    SATUAN=="MOBIL" present. Across the two real document variants
    seen so far, that's the one signal that's actually constant; NO
    is not. Dropping the NO requirement entirely rather than trying
    to special-case both shapes — "unit is literally MOBIL" is
    already a strong, near-unambiguous signal on its own (no genuine
    line item's unit would ever legitimately read "MOBIL").
    """
    if satuan_cell is None:
        return False
    return str(satuan_cell).strip().upper() == "MOBIL"


def _split_vehicle_name(raw_name: str):
    """
    "HYUNDAY TUCSON (9-XXXI)" -> ("HYUNDAY TUCSON", "9-XXXI"). Falls
    back to the whole string as the model name with an empty fleet
    code if no parenthesized code is found — surfaced as a review-
    time gap (the reviewer can fill in a plate number by hand) rather
    than a hard parse failure, since one malformed vehicle name
    shouldn't block importing every other vehicle in the same file.
    """
    match = FLEET_CODE_RE.search(raw_name)
    if not match:
        return raw_name.strip(), ""
    fleet_code = match.group(1).strip()
    model_name = raw_name[: match.start()].strip()
    return model_name, fleet_code


def parse_hps_workbook(file_obj) -> ParsedContract:
    """
    Entry point. Locates the header row by searching for the literal
    column labels rather than assuming a fixed row number — the real
    document has a title block (and, per Chris, sometimes a redacted
    reference line) above the table of unknown exact height, so
    anchoring to a fixed row number would break the moment that
    header block's height changes even slightly between contracts.
    """
    workbook = openpyxl.load_workbook(file_obj, data_only=True)
    sheet = workbook.active

    header_row_idx = None
    for row_idx, row in enumerate(sheet.iter_rows(min_row=1, max_row=30), start=1):
        values = [str(c.value).strip().upper() if c.value is not None else "" for c in row]
        if "ITEM PEKERJAAN" in values and "SATUAN" in values:
            header_row_idx = row_idx
            break

    if header_row_idx is None:
        raise ContractParseError(
            "Could not locate the header row (expected columns including "
            "'ITEM PEKERJAAN' and 'SATUAN') in the first 30 rows."
        )

    header_cells = [str(c.value).strip().upper() if c.value is not None else "" for c in sheet[header_row_idx]]
    try:
        col = {
            "no":      header_cells.index("NO"),
            "item":    header_cells.index("ITEM PEKERJAAN"),
            "vol":     header_cells.index("VOL"),
            "satuan":  header_cells.index("SATUAN"),
            "harga":   header_cells.index("HARGA SATUAN"),
            "jumlah":  header_cells.index("JUMLAH"),
        }
    except ValueError as e:
        raise ContractParseError(f"Expected column not found in header row: {e}")

    vehicle_groups: List[ParsedVehicleGroup] = []
    current_group: Optional[ParsedVehicleGroup] = None
    document_total: Optional[Decimal] = None

    for row in sheet.iter_rows(min_row=header_row_idx + 1):
        no_val     = row[col["no"]].value
        item_val   = row[col["item"]].value
        satuan_val = row[col["satuan"]].value

        if item_val is None and no_val is None:
            continue  # blank spacer row — the real document has these between printed pages

        item_text = str(item_val).strip() if item_val is not None else ""

        if item_text.upper() == "TOTAL KESELURUHAN":
            jumlah_val = row[col["jumlah"]].value
            if jumlah_val is not None:
                document_total = parse_rupiah(jumlah_val)
            continue

        if _is_group_header_row(satuan_val):
            if current_group is not None:
                vehicle_groups.append(current_group)
            model_name, fleet_code = _split_vehicle_name(item_text)
            jumlah_val = row[col["jumlah"]].value
            current_group = ParsedVehicleGroup(
                # NO is often blank on the real group-header row (see
                # _is_group_header_row's own docstring) — group_label
                # is purely informational and never used for matching
                # (fleet_code is), so a blank one is harmless.
                group_label=str(no_val).strip() if no_val is not None else "",
                raw_name=item_text,
                vehicle_model=model_name,
                fleet_code=fleet_code,
                allocated_budget=parse_rupiah(jumlah_val) if jumlah_val is not None else Decimal("0"),
            )
            continue

        if current_group is None:
            # A line-item-shaped row appeared before any vehicle
            # group header was ever found — the template assumption
            # doesn't hold for this particular file. Fail loudly
            # rather than silently attaching orphaned line items to
            # nothing.
            raise ContractParseError(
                f"Line item row found before any vehicle group header: {item_text!r}"
            )

        try:
            row_no = int(str(no_val).strip())
        except (TypeError, ValueError):
            continue  # not a real numbered row — skip rather than guess at meaning

        current_group.line_items.append(ParsedLineItem(
            row_no=row_no,
            description=item_text,
            volume=parse_rupiah(row[col["vol"]].value) if row[col["vol"]].value is not None else Decimal("0"),
            unit=str(row[col["satuan"]].value).strip() if row[col["satuan"]].value is not None else "",
            unit_price=parse_rupiah(row[col["harga"]].value),
            subtotal=parse_rupiah(row[col["jumlah"]].value),
        ))

    if current_group is not None:
        vehicle_groups.append(current_group)

    if not vehicle_groups:
        raise ContractParseError("No vehicle groups were found in this workbook.")

    return ParsedContract(vehicle_groups=vehicle_groups, document_total=document_total)


def diff_against_contract(parsed: ParsedContract, contract) -> dict:
    """
    Compares a fresh parse against the CURRENT ACTIVE state of a real
    Contract's line items. Matching key is (fleet_code,
    source_row_no) — see ContractLineItem's own docstring for why
    description text isn't trustworthy as a match key across a
    revision.

    Returns a plain dict, JSON-serializable as-is for
    ContractImport.parsed_diff — every Decimal is stringified here so
    json.dumps() never has to deal with a raw Decimal at all.
    """
    from apps.service.models import Vehicle  # local import, same

    # reasoning as the ContractVehicle import just below — parsing.py
    # stays importable without Django's ORM configured for its own
    # pure-parsing tests, this only gets pulled in when the diff
    # function (which always needs the ORM anyway) actually runs.
    from .models import ContractVehicle

    existing_vehicles = {
        cv.vehicle.plate_number: cv
        for cv in ContractVehicle.objects.filter(contract=contract).select_related("vehicle")
    }

    diff = {
        "added_vehicles": [],
        "added_items":    [],
        "changed_items":  [],
        "removed_items":  [],
        "unchanged_count": 0,
    }

    seen_positions = set()  # (fleet_code, row_no) pairs present in this parse

    for group in parsed.vehicle_groups:
        existing_cv = existing_vehicles.get(group.fleet_code)

        if existing_cv is None:
            # New to THIS contract — but not necessarily new to the
            # org. The exact scenario ContractVehicle was built as a
            # separate join table for in the first place: the same
            # real fleet vehicle reappearing in a later fiscal year's
            # contract. Vehicle.plate_number is unique per
            # organization (not per contract), so checking here,
            # before ever proposing a "create a new Vehicle" action,
            # is what actually lets that reuse work — surfacing
            # existing_vehicle_id lets the review screen skip asking
            # for manufacture_year again (the existing Vehicle
            # already has one) and lets apply() link to it instead of
            # colliding with the DB's own unique constraint.
            reusable_vehicle = Vehicle.objects.filter(
                organization=contract.organization, plate_number=group.fleet_code,
            ).first()
            diff["added_vehicles"].append({
                "fleet_code": group.fleet_code,
                "vehicle_model": group.vehicle_model,
                "allocated_budget": str(group.allocated_budget),
                "existing_vehicle_id": str(reusable_vehicle.id) if reusable_vehicle else None,
                "existing_vehicle_model": reusable_vehicle.model if reusable_vehicle else None,
                "line_items": [
                    {
                        "row_no": li.row_no, "description": li.description,
                        "volume": str(li.volume), "unit": li.unit,
                        "unit_price": str(li.unit_price), "subtotal": str(li.subtotal),
                    }
                    for li in group.line_items
                ],
            })
            continue

        existing_items = {
            li.source_row_no: li
            for li in existing_cv.line_items.filter(status="ACTIVE")
        }

        for li in group.line_items:
            seen_positions.add((group.fleet_code, li.row_no))
            existing_li = existing_items.get(li.row_no)

            if existing_li is None:
                diff["added_items"].append({
                    "fleet_code": group.fleet_code, "row_no": li.row_no,
                    "description": li.description, "volume": str(li.volume),
                    "unit": li.unit, "unit_price": str(li.unit_price),
                    "subtotal": str(li.subtotal),
                })
                continue

            changed = (
                existing_li.description != li.description
                or existing_li.volume != li.volume
                or existing_li.unit != li.unit
                or existing_li.unit_price != li.unit_price
                or existing_li.subtotal != li.subtotal
            )
            if changed:
                diff["changed_items"].append({
                    "fleet_code": group.fleet_code, "row_no": li.row_no,
                    "old": {
                        "description": existing_li.description, "volume": str(existing_li.volume),
                        "unit": existing_li.unit, "unit_price": str(existing_li.unit_price),
                        "subtotal": str(existing_li.subtotal),
                    },
                    "new": {
                        "description": li.description, "volume": str(li.volume),
                        "unit": li.unit, "unit_price": str(li.unit_price),
                        "subtotal": str(li.subtotal),
                    },
                })
            else:
                diff["unchanged_count"] += 1

    for fleet_code, cv in existing_vehicles.items():
        for li in cv.line_items.filter(status="ACTIVE"):
            if (fleet_code, li.source_row_no) not in seen_positions:
                diff["removed_items"].append({
                    "fleet_code": fleet_code, "row_no": li.source_row_no,
                    "description": li.description,
                })

    return diff
